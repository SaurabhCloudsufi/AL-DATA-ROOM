#!/usr/bin/env python3
"""Link the Excel register's Plot IDs to the published chart gallery.

Run this once GitHub Pages is live and the real URL is known. Nothing here
guesses a username or a URL - the base address is supplied on the command line.

What it does:
  02_Plot_Index      fills Chart_link with the gallery anchor for each plot,
                     as a clickable link reading "View chart"
  01_Dataset_Register turns the Plots column into a link to the first plot
                     listed for that dataset, so a reader can jump from a
                     dataset straight into its evidence

Usage:
    python build/update_excel_links.py <workbook.xlsx> <base-url>

Example:
    python build/update_excel_links.py register.xlsx \\
        https://your-username.github.io/ai-data-room
"""
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

DOMAIN = "inference-tokens"
LINK_FONT = Font(name="Arial", size=9, color="0563C1", underline="single")


def main(xlsx: str, base: str) -> None:
    base = base.rstrip("/")
    gallery = f"{base}/{DOMAIN}/"
    path = Path(xlsx)
    wb = openpyxl.load_workbook(path, rich_text=True)

    # ---- 02_Plot_Index: one anchor per plot --------------------------
    ws = wb["02_Plot_Index"]
    linked = 0
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(row=r, column=1).value
        if not pid:
            continue
        cell = ws.cell(row=r, column=6)          # Chart_link
        cell.value = "View chart"
        cell.hyperlink = f"{gallery}#{pid}"
        cell.font = LINK_FONT
        linked += 1

    # ---- 01_Dataset_Register: jump into the evidence -----------------
    ws1 = wb["01_Dataset_Register"]
    jumped = 0
    for r in range(2, ws1.max_row + 1):
        cell = ws1.cell(row=r, column=6)         # Plots/graph/charts
        m = re.search(r"P-\d{2}", str(cell.value or ""))
        if not m:
            continue
        cell.hyperlink = f"{gallery}#{m.group(0)}"
        cell.font = LINK_FONT
        jumped += 1

    # ---- 04_Read_Me: record where the gallery lives ------------------
    ws4 = wb["04_Read_Me"]
    row = ws4.max_row + 2
    a = ws4.cell(row=row, column=2, value="Chart gallery")
    a.font = Font(name="Arial", size=9, bold=True)
    b = ws4.cell(row=row, column=3, value=gallery)
    b.hyperlink = gallery
    b.font = LINK_FONT

    wb.save(path)
    print(f"gallery base : {gallery}")
    print(f"02_Plot_Index: {linked} plot links written")
    print(f"01_Dataset_Register: {jumped} dataset rows linked to their first chart")
    print(f"saved {path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    main(sys.argv[1], sys.argv[2])
