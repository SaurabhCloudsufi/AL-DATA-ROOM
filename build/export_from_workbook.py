#!/usr/bin/env python3
"""Export the Plot Index and Company Disclosures from the Excel register into CSVs.

The Excel workbook remains the single source of truth for the plot inventory.
This script is the one-way bridge from the workbook into the site build, so the
gallery can never contain a chart that is not in the Plot Index.

Usage:
    python build/export_from_workbook.py path/to/Inference_Tokens_Dataset_Register_Client.xlsx
"""
import csv
import re
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / "build"

PLOT_COLS = ["Plot_ID", "Chart_title", "Datasets_used", "What_the_chart_shows",
             "Methodology_reference", "Chart_link", "Status"]

DISC_COLS = ["Company", "Disclosure_date", "Metric", "Value_as_disclosed",
             "Unit_as_disclosed", "Normalized_value", "Normalized_unit", "Scope",
             "Product_or_surface", "Disclosure_context", "Source_name", "Source_link",
             "Source_type", "Methodology_reference", "Plot_ID", "Notes"]


def main(xlsx: str) -> None:
    wb = openpyxl.load_workbook(xlsx)

    # ---- Plot Index -----------------------------------------------------
    ws = wb["02_Plot_Index"]
    plots = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, 8)]
        if row[0]:
            plots.append(row)

    # ---- owning dataset + source, taken from the Dataset Register -------
    ws1 = wb["01_Dataset_Register"]
    owner = {}
    for r in range(2, ws1.max_row + 1):
        dataset = ws1.cell(row=r, column=2).value or ""
        src_cell = ws1.cell(row=r, column=4)
        src_name = (src_cell.value or "").split("\n")[0]
        src_url = src_cell.hyperlink.target if src_cell.hyperlink else ""
        for m in re.finditer(r"P-\d{2}", ws1.cell(row=r, column=6).value or ""):
            owner.setdefault(m.group(0), (dataset, src_name, src_url))

    with (OUT / "plot_index.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(PLOT_COLS + ["Owning_dataset", "Source_name", "Source_url"])
        for p in plots:
            ds, sn, su = owner.get(p[0], ("", "", ""))
            w.writerow(p + [ds, sn, su])

    # ---- Company Disclosures -------------------------------------------
    ws3 = wb["03_Company_Disclosures"]
    with (OUT / "company_disclosures.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(DISC_COLS)
        for r in range(2, ws3.max_row + 1):
            row = [ws3.cell(row=r, column=c).value for c in range(1, 17)]
            link = ws3.cell(row=r, column=12)
            row[11] = link.hyperlink.target if link.hyperlink else ""
            if row[0]:
                w.writerow(row)

    # a public copy travels with the site so the notebook is reproducible
    pub = REPO / "inference-tokens" / "data" / "company_disclosures.csv"
    pub.write_text((OUT / "company_disclosures.csv").read_text(encoding="utf-8"),
                   encoding="utf-8")

    print(f"plot_index.csv          {len(plots)} plots")
    print(f"company_disclosures.csv written to build/ and inference-tokens/data/")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("usage: export_from_workbook.py <workbook.xlsx>")
    main(sys.argv[1])
